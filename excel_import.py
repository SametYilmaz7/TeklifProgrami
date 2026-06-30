"""
excel_import.py
~~~~~~~~~~~~~~~
Excel dosyasından toplu ürün içe aktarma modülü.

Desteklenen sütunlar:
  - Ürün Kodu  (zorunlu)
  - Ürün Adı   (zorunlu)
  - Link       (opsiyonel)

Resim eşleştirme:
  images_dir verilmişse ürün kodunu dosya adı olarak arar.
  Örn: PRD-001 → PRD-001.png / .jpg / .jpeg / .webp / .bmp / .gif
"""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

IMAGE_EXTENSIONS = [".png", ".jpg", ".jpeg", ".webp", ".bmp", ".gif"]

_COLUMN_ALIASES = {
    "ürünkodu": "product_code",
    "urunkodu": "product_code",
    "kod": "product_code",
    "code": "product_code",
    "productcode": "product_code",
    "ürünadı": "name",
    "urunadi": "name",
    "ad": "name",
    "adi": "name",
    "name": "name",
    "ürünadi": "name",
    "link": "product_url",
    "url": "product_url",
    "producturl": "product_url",
    "teknikbilgi": "product_url",
    "tekniklink": "product_url",
    "teknikbilgilinki": "product_url",
}


def _normalize(text: str) -> str:
    return (
        str(text)
        .lower()
        .replace(" ", "")
        .replace("-", "")
        .replace("_", "")
        .replace("(", "")
        .replace(")", "")
        .strip()
    )


@dataclass
class ImportResult:
    added: int = 0
    skipped_duplicate: list[str] = field(default_factory=list)
    skipped_missing_fields: list[int] = field(default_factory=list)
    skipped_image_not_found: list[str] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)

    def summary(self) -> str:
        lines = [f"✅ {self.added} ürün başarıyla eklendi."]
        if self.skipped_duplicate:
            codes = ", ".join(self.skipped_duplicate)
            lines.append(
                f"⚠️  {len(self.skipped_duplicate)} ürün zaten mevcut (atlandı): {codes}"
            )
        if self.skipped_missing_fields:
            rows = ", ".join(str(r) for r in self.skipped_missing_fields)
            lines.append(
                f"⚠️  {len(self.skipped_missing_fields)} satır eksik alan nedeniyle atlandı (satırlar: {rows})"
            )
        if self.skipped_image_not_found:
            lines.append(
                f"ℹ️  {len(self.skipped_image_not_found)} ürünün resmi klasörde bulunamadı "
                f"(ürün yine de eklendi)"
            )
        if self.errors:
            lines.append("❌ Hatalar:\n" + "\n".join(self.errors))
        return "\n".join(lines)


def _find_image_by_code(product_code: str, images_dir: Path) -> str:
    """Ürün kodunu dosya adı olarak kullanarak resmi arar."""
    for ext in IMAGE_EXTENSIONS:
        candidate = images_dir / (product_code + ext)
        if candidate.exists():
            return str(candidate.resolve())
    return ""


def import_from_excel(
    excel_path: str,
    images_dir: Optional[str] = None,
) -> ImportResult:
    try:
        import openpyxl
    except ImportError:
        result = ImportResult()
        result.errors.append(
            "openpyxl kurulu değil. Lütfen 'pip install openpyxl' komutunu çalıştırın."
        )
        return result

    from database import add_product, DuplicateProductError, DatabaseError

    result = ImportResult()
    img_dir = Path(images_dir).resolve() if images_dir else None

    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
    except Exception as e:
        result.errors.append(f"Excel dosyası açılamadı: {e}")
        return result

    ws = wb.active
    header_row = None
    col_map: dict[str, int] = {}

    for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
        if any(cell is not None for cell in row):
            for col_idx, cell in enumerate(row):
                if cell is None:
                    continue
                norm = _normalize(str(cell))
                field_name = _COLUMN_ALIASES.get(norm)
                if field_name:
                    col_map[field_name] = col_idx
            header_row = row_idx
            break

    if header_row is None or "product_code" not in col_map or "name" not in col_map:
        result.errors.append(
            "Başlık satırı bulunamadı veya zorunlu sütunlar eksik.\n"
            "Excel'de en az 'Ürün Kodu' ve 'Ürün Adı' sütunları bulunmalıdır."
        )
        return result

    for row_idx, row in enumerate(
        ws.iter_rows(min_row=header_row + 2, values_only=True),
        start=header_row + 2,
    ):
        if all(cell is None or str(cell).strip() == "" for cell in row):
            continue

        def get(f: str) -> str:
            idx = col_map.get(f)
            if idx is None or idx >= len(row):
                return ""
            val = row[idx]
            return str(val).strip() if val is not None else ""

        product_code = get("product_code")
        name = get("name")
        product_url = get("product_url")

        if not product_code or not name:
            result.skipped_missing_fields.append(row_idx)
            continue

        if img_dir:
            image_path = _find_image_by_code(product_code, img_dir)
            if not image_path:
                result.skipped_image_not_found.append(product_code)
        else:
            image_path = ""

        try:
            add_product(product_code, name, image_path, product_url)
            result.added += 1
        except DuplicateProductError:
            result.skipped_duplicate.append(product_code)
        except DatabaseError as e:
            result.errors.append(f"Satır {row_idx} ({product_code}): {e}")

    return result


def create_template(save_path: str) -> None:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ürünler"

    HEADER_BG = "1A1A2E"
    HEADER_FG = "FFFFFF"
    EXAMPLE_BG = "F0F4FF"
    BORDER_CLR = "CCCCCC"
    NOTE_BG = "FFF9C4"

    thin = Side(style="thin", color=BORDER_CLR)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ["Ürün Kodu", "Ürün Adı", "Link"]
    col_widths = [20, 40, 55]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(name="Arial", bold=True, color=HEADER_FG, size=11)
        cell.fill = PatternFill("solid", fgColor=HEADER_BG)
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 30

    ws.freeze_panes = "A2"
    wb.save(save_path)
